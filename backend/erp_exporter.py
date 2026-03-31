"""
AuraGrade — ERP Export Module & Digital Signature Layer
=======================================================
Converts validated grades into official university ledger formats (CSV / Excel)
and generates SHA-256 integrity hashes for tamper-proof compliance.

This is the final administrative closure step: once the CoE finalises marks,
a digitally signed ledger is generated that can be sent to the university ERP
system. The hash acts as a legal seal — any modification to the file after
export will produce a mismatched hash.

Exports
-------
- generate_university_ledger()  → creates CSV/XLSX in-memory, returns bytes + hash
- generate_integrity_hash()     → SHA-256 hash of any byte content
- generate_ledger_preview()     → lightweight JSON preview (no file generation)
"""

from __future__ import annotations

import hashlib
import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ---------------------------------------------------------------------------
#  Digital Signature — SHA-256 Integrity Hash
# ---------------------------------------------------------------------------

def generate_integrity_hash(data: bytes) -> str:
    """
    Generate a SHA-256 digital fingerprint for arbitrary byte content.

    This acts as a tamper-proof seal: if even a single byte is changed
    in the exported ledger, the hash will no longer match.

    Parameters
    ----------
    data : bytes
        The raw content to hash (CSV bytes, XLSX bytes, etc.)

    Returns
    -------
    str
        64-character lowercase hex SHA-256 digest.
    """
    sha256 = hashlib.sha256()
    # Process in 4 KB chunks for memory efficiency with large datasets
    offset = 0
    chunk_size = 4096
    while offset < len(data):
        sha256.update(data[offset : offset + chunk_size])
        offset += chunk_size
    return sha256.hexdigest()


# ---------------------------------------------------------------------------
#  Ledger Generation — CSV & Excel
# ---------------------------------------------------------------------------

# Shared openpyxl style constants
_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=14)
_SUBHEADER_FONT = Font(size=10, italic=True, color="64748B")
_COL_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_COL_HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
_COL_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SEAL_LABEL_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
_SEAL_LABEL_FONT = Font(bold=True, size=10, color="0F172A")
_SEAL_HASH_FONT = Font(name="Courier New", size=9, color="0F172A")
_SEAL_HASH_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    top=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)
_SENTINEL_CRITICAL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_SENTINEL_CRITICAL_FONT = Font(bold=True, color="DC2626", size=9)
_SENTINEL_WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_SENTINEL_WARNING_FONT = Font(bold=True, color="D97706", size=9)
_SENTINEL_CLEAR_FONT = Font(color="16A34A", size=9)
_DATA_FONT = Font(size=9, color="1E293B")
_DATA_ALIGN = Alignment(horizontal="center", vertical="center")
_ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")


def _apply_institutional_styling(
    worksheet,
    title_text: str,
    subtitle_text: str,
    df: pd.DataFrame,
    data_start_row: int,
    sha256_hash: str,
    total_cols: int,
    sentinel_col_indices: list[int] | None = None,
):
    """
    Apply professional institutional formatting to an openpyxl worksheet.

    Parameters
    ----------
    worksheet : openpyxl Worksheet
    title_text : str — Title line (row 1).
    subtitle_text : str — Subtitle line (row 2).
    df : pd.DataFrame — The data written starting at data_start_row.
    data_start_row : int — 1-based row where the header + data begin.
    sha256_hash : str — 64-char hex digest to embed as seal.
    total_cols : int — Number of columns.
    sentinel_col_indices : list[int] — 0-based column indices for sentinel
        status cells to apply conditional colouring.
    """
    last_col_letter = chr(ord("A") + min(total_cols - 1, 25))  # up to Z

    # ── Row 1: Title ──
    worksheet.merge_cells(f"A1:{last_col_letter}1")
    cell_title = worksheet["A1"]
    cell_title.value = title_text
    cell_title.font = _HEADER_FONT
    cell_title.fill = _HEADER_FILL
    cell_title.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 32

    # ── Row 2: Subtitle ──
    worksheet.merge_cells(f"A2:{last_col_letter}2")
    cell_sub = worksheet["A2"]
    cell_sub.value = subtitle_text
    cell_sub.font = _SUBHEADER_FONT
    cell_sub.alignment = Alignment(horizontal="left")
    worksheet.row_dimensions[2].height = 20

    # ── Row 3: Blank separator ──
    worksheet.row_dimensions[3].height = 6

    # ── Row data_start_row (column headers) ──
    header_row = data_start_row
    for col_idx in range(1, total_cols + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.fill = _COL_HEADER_FILL
        cell.font = _COL_HEADER_FONT
        cell.alignment = _COL_HEADER_ALIGN
        cell.border = _THIN_BORDER
    worksheet.row_dimensions[header_row].height = 24

    # ── Data rows ──
    num_data_rows = len(df)
    for row_offset in range(num_data_rows):
        row_num = header_row + 1 + row_offset
        for col_idx in range(1, total_cols + 1):
            cell = worksheet.cell(row=row_num, column=col_idx)
            cell.font = _DATA_FONT
            cell.alignment = _DATA_ALIGN
            cell.border = Border(
                bottom=Side(border_style="hair", color="E2E8F0"),
            )
            # Alternating row fill
            if row_offset % 2 == 1:
                cell.fill = _ALT_ROW_FILL

        # Sentinel conditional colouring
        if sentinel_col_indices:
            for s_col in sentinel_col_indices:
                cell = worksheet.cell(row=row_num, column=s_col + 1)  # 1-based
                val = str(cell.value or "").strip()
                if val == "Critical":
                    cell.fill = _SENTINEL_CRITICAL_FILL
                    cell.font = _SENTINEL_CRITICAL_FONT
                elif val == "Warning":
                    cell.fill = _SENTINEL_WARNING_FILL
                    cell.font = _SENTINEL_WARNING_FONT
                elif val == "Clear":
                    cell.font = _SENTINEL_CLEAR_FONT

    # ── Auto-fit column widths (approximate) ──
    for col_idx in range(1, total_cols + 1):
        max_len = 0
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        for row_num in range(header_row, header_row + num_data_rows + 1):
            val = str(worksheet.cell(row=row_num, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)

    # ── Seal Footer ──
    seal_label_row = header_row + num_data_rows + 2
    seal_hash_row = seal_label_row + 1
    seal_note_row = seal_hash_row + 1

    worksheet.merge_cells(f"A{seal_label_row}:{last_col_letter}{seal_label_row}")
    label_cell = worksheet[f"A{seal_label_row}"]
    label_cell.value = "DIGITAL INTEGRITY SEAL (SHA-256)"
    label_cell.font = _SEAL_LABEL_FONT
    label_cell.fill = _SEAL_LABEL_FILL
    label_cell.alignment = Alignment(horizontal="center")
    label_cell.border = _THIN_BORDER
    worksheet.row_dimensions[seal_label_row].height = 22

    worksheet.merge_cells(f"A{seal_hash_row}:{last_col_letter}{seal_hash_row}")
    hash_cell = worksheet[f"A{seal_hash_row}"]
    hash_cell.value = sha256_hash
    hash_cell.font = _SEAL_HASH_FONT
    hash_cell.alignment = _SEAL_HASH_ALIGN
    hash_cell.border = _THIN_BORDER
    worksheet.row_dimensions[seal_hash_row].height = 20

    worksheet.merge_cells(f"A{seal_note_row}:{last_col_letter}{seal_note_row}")
    note_cell = worksheet[f"A{seal_note_row}"]
    note_cell.value = (
        "Any modification to this file will invalidate the seal above. "
        "Verify at AuraGrade Portal → Digital Seal → Upload & Verify."
    )
    note_cell.font = Font(size=8, italic=True, color="94A3B8")
    note_cell.alignment = Alignment(horizontal="center")

    # Freeze header row
    worksheet.freeze_panes = f"A{header_row + 1}"

def _build_ledger_dataframe(grades_data: list[dict]) -> pd.DataFrame:
    """
    Flatten Supabase grade rows into a university-standard ledger DataFrame.

    Expected input: list of grade rows with nested students() and assessments().
    """
    rows = []
    for entry in grades_data:
        student = entry.get("students") or {}
        assessment = entry.get("assessments") or {}
        rows.append({
            "Sl_No": len(rows) + 1,
            "Register_Number": student.get("reg_no", "N/A"),
            "Student_Name": student.get("name", "N/A"),
            "Subject": assessment.get("subject", "N/A"),
            "Assessment": assessment.get("title", "N/A"),
            "Internal_Marks": round(entry.get("ai_score", 0), 2),
            "Confidence": round(entry.get("confidence", 0) * 100, 1),
            "Verification_Status": entry.get("prof_status", "Pending"),
            "Flagged": "Yes" if entry.get("is_flagged") else "No",
            "Graded_At": entry.get("graded_at", ""),
            "Reviewed_At": entry.get("reviewed_at", "") or "",
            "Export_Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    return pd.DataFrame(rows)


def generate_university_ledger(
    grades_data: list[dict],
    assessment_id: str,
    fmt: str = "csv",
) -> dict:
    """
    Generate a university ERP-compatible ledger from validated grades.

    Parameters
    ----------
    grades_data : list[dict]
        Supabase grade rows (with nested students + assessments).
    assessment_id : str
        UUID of the assessment (used in the filename).
    fmt : str
        'csv' or 'xlsx'.

    Returns
    -------
    dict with keys:
        filename : str          — generated filename
        content  : bytes        — raw file bytes (ready for download)
        sha256   : str          — integrity hash
        records  : int          — number of student records
        format   : str          — 'csv' or 'xlsx'
    """
    if not grades_data:
        return {
            "filename": None,
            "content": b"",
            "sha256": "",
            "records": 0,
            "format": fmt,
            "error": "No approved/audited grades found for this assessment.",
        }

    df = _build_ledger_dataframe(grades_data)
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_id = assessment_id[:8]

    if fmt == "xlsx":
        # First pass: generate raw data to compute hash
        buf_raw = io.BytesIO()
        with pd.ExcelWriter(buf_raw, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Marks Ledger", startrow=3)
        raw_bytes = buf_raw.getvalue()
        sha256 = generate_integrity_hash(raw_bytes)

        # Second pass: styled workbook with seal
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Marks Ledger", startrow=3)
            ws = writer.sheets["Marks Ledger"]

            # Determine subject from first entry
            first_assessment = (grades_data[0].get("assessments") or {}) if grades_data else {}
            subject = first_assessment.get("subject", "")
            title = first_assessment.get("title", "")
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            _apply_institutional_styling(
                worksheet=ws,
                title_text=f"AURAGRADE MARKS LEDGER — {subject}".upper(),
                subtitle_text=f"Assessment: {title} | ID: {assessment_id[:8]}… | Exported: {timestamp}",
                df=df,
                data_start_row=4,  # row 4 is the header (startrow=3 is 0-based)
                sha256_hash=sha256,
                total_cols=len(df.columns),
            )

        content = buf.getvalue()
        filename = f"LEDGER_{short_id}_{date_str}.xlsx"
    else:
        content = df.to_csv(index=False).encode("utf-8")
        filename = f"LEDGER_{short_id}_{date_str}.csv"

    sha256 = generate_integrity_hash(content)

    return {
        "filename": filename,
        "content": content,
        "sha256": sha256,
        "records": len(df),
        "format": fmt,
    }


def generate_ledger_preview(grades_data: list[dict], limit: int = 10) -> dict:
    """
    Return a lightweight JSON preview of what the ledger will contain.
    Used by the "Preview Ledger" button in the UI before final export.

    Returns
    -------
    dict with keys:
        total_records : int
        preview       : list[dict]  — first N rows as dicts
        columns       : list[str]   — column headers
    """
    if not grades_data:
        return {"total_records": 0, "preview": [], "columns": []}

    df = _build_ledger_dataframe(grades_data)
    return {
        "total_records": len(df),
        "preview": df.head(limit).to_dict(orient="records"),
        "columns": list(df.columns),
    }


# ---------------------------------------------------------------------------
#  Institutional Ledger — Enhanced with Sentinel + Digital Seal
# ---------------------------------------------------------------------------

def _build_institutional_dataframe(
    grades_data: list[dict],
    sentinel_flags: dict[str, dict] | None = None,
    sha256_hash: str = "",
) -> pd.DataFrame:
    """
    Build a comprehensive institutional ledger that includes:
      - Standard student marks & verification status
      - Similarity Sentinel flag status per student
      - SHA-256 digital seal reference

    Parameters
    ----------
    grades_data : list[dict]
        Supabase grade rows (with nested students + assessments).
    sentinel_flags : dict mapping student_id → {status, similarity, peer}
        Pre-computed collusion flags (from scan_assessment_collusion).
    sha256_hash : str
        The digital seal hash to stamp on every row.
    """
    rows = []
    for entry in grades_data:
        student = entry.get("students") or {}
        assessment = entry.get("assessments") or {}
        student_id = entry.get("student_id", "")

        # Sentinel lookup
        sentinel = (sentinel_flags or {}).get(student_id, {})
        sentinel_status = sentinel.get("status", "Clear")
        sentinel_similarity = sentinel.get("similarity", "")
        sentinel_peer = sentinel.get("peer", "")

        rows.append({
            "Sl_No": len(rows) + 1,
            "Register_Number": student.get("reg_no", "N/A"),
            "Student_Name": student.get("name", "N/A"),
            "Subject": assessment.get("subject", "N/A"),
            "Assessment": assessment.get("title", "N/A"),
            "Internal_Marks": round(entry.get("ai_score", 0), 2),
            "Confidence_Pct": round(entry.get("confidence", 0) * 100, 1),
            "Verification_Status": entry.get("prof_status", "Pending"),
            "AI_Flagged": "Yes" if entry.get("is_flagged") else "No",
            "Sentinel_Status": sentinel_status,
            "Sentinel_Similarity": sentinel_similarity,
            "Sentinel_Peer": sentinel_peer,
            "Graded_At": entry.get("graded_at", ""),
            "Reviewed_At": entry.get("reviewed_at", "") or "",
            "Digital_Seal_SHA256": sha256_hash[:16] + "…" if sha256_hash else "—",
            "Export_Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    return pd.DataFrame(rows)


def generate_institutional_ledger(
    grades_data: list[dict],
    assessment_id: str,
    sentinel_flags: dict[str, dict] | None = None,
    fmt: str = "csv",
) -> dict:
    """
    Generate the comprehensive institutional ledger with sentinel flags
    and SHA-256 digital seal.  XLSX exports include:

    - Institutional header with subject & timestamp
    - Styled column headers (dark fill, white text)
    - Conditional colouring for Sentinel columns (Critical=red, Warning=amber)
    - Alternating row stripes
    - SHA-256 tamper-proof seal footer row

    Returns
    -------
    dict with keys:
        filename : str
        content  : bytes
        sha256   : str
        records  : int
        format   : str
    """
    if not grades_data:
        return {
            "filename": None,
            "content": b"",
            "sha256": "",
            "records": 0,
            "format": fmt,
            "error": "No approved/audited grades found for this assessment.",
        }

    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_id = assessment_id[:8]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Determine subject/title from first entry
    first_assessment = (grades_data[0].get("assessments") or {}) if grades_data else {}
    subject = first_assessment.get("subject", "")
    title = first_assessment.get("title", "")

    if fmt == "xlsx":
        # First pass — build without the seal column to compute hash
        df_raw = _build_institutional_dataframe(grades_data, sentinel_flags, sha256_hash="")
        buf_raw = io.BytesIO()
        with pd.ExcelWriter(buf_raw, engine="openpyxl") as writer:
            df_raw.to_excel(writer, index=False, sheet_name="Institutional Ledger", startrow=3)
        sha256 = generate_integrity_hash(buf_raw.getvalue())

        # Second pass — stamp the seal into the data AND the footer
        df_final = _build_institutional_dataframe(grades_data, sentinel_flags, sha256_hash=sha256)

        # Identify sentinel column indices (0-based in the DataFrame)
        sentinel_col_names = {"Sentinel_Status", "Sentinel_Similarity", "Sentinel_Peer"}
        sentinel_col_indices = [
            i for i, col in enumerate(df_final.columns) if col in sentinel_col_names
        ]

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_final.to_excel(writer, index=False, sheet_name="Institutional Ledger", startrow=3)
            ws = writer.sheets["Institutional Ledger"]

            _apply_institutional_styling(
                worksheet=ws,
                title_text=f"AURAGRADE INSTITUTIONAL LEDGER — {subject}".upper(),
                subtitle_text=(
                    f"Assessment: {title} | ID: {assessment_id[:8]}… | "
                    f"Exported: {timestamp} | "
                    f"Sentinel Flags: {len(sentinel_flags or {})}"
                ),
                df=df_final,
                data_start_row=4,
                sha256_hash=sha256,
                total_cols=len(df_final.columns),
                sentinel_col_indices=sentinel_col_indices,
            )

        content = buf.getvalue()
        filename = f"INST_LEDGER_{short_id}_{date_str}.xlsx"

        # Final hash of the fully styled workbook
        sha256 = generate_integrity_hash(content)

    else:
        # CSV — flat export, no styling possible
        df_raw = _build_institutional_dataframe(grades_data, sentinel_flags, sha256_hash="")
        raw_csv = df_raw.to_csv(index=False).encode("utf-8")
        sha256 = generate_integrity_hash(raw_csv)

        df_final = _build_institutional_dataframe(grades_data, sentinel_flags, sha256_hash=sha256)
        content = df_final.to_csv(index=False).encode("utf-8")
        filename = f"INST_LEDGER_{short_id}_{date_str}.csv"

        # Re-hash with seal column included
        sha256 = generate_integrity_hash(content)

    return {
        "filename": filename,
        "content": content,
        "sha256": sha256,
        "records": len(df_final),
        "format": fmt,
    }

